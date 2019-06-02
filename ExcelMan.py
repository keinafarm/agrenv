# -*- coding: utf-8 -*-

import openpyxl
import re

class   ExcelMan:
    """
    Excel File 管理クラス

    openpyxlの使い方
    https://www.python-izm.com/third_party/excel/openpyxl/
    https://tonari-it.com/python-openpyxl-beginner/
    https://openpyxl.readthedocs.io/en/stable/
    """
    __fileName = None       # 現在開いているブックのファイル名
    __book = None           # 現在開いているブック
    __sheets = None         # 現在フォーカスしているシート

    def __init__(self, file):
        """
        コンストラクタ
        :param file: Excelファイル名
        """
        self.__fileName = file

    def openBook(self):
        """
        ブックをオープンする
        """
        self.__book = openpyxl.load_workbook(self.__fileName)

    def getSheetList(self):
        """
        シート名一覧を得る
        :return:    シート名一覧のリスト
        """
        return self.__book.sheetnames

    def openSheet(self, sheetIndex):
        """
        指定したシートをオープンする
        :param sheetIndex: シート番号
        """
        self.__sheet =  self.__book.worksheets[sheetIndex]

    def getCell(self, row, col):
        """
        指定したセルを得る
        :param row: 行番号
        :param col: 列番号
        :return: セルオブジェクト
        """
        return  self.__sheet.cell( row=row, column=col )

    def getCellValue(self, row, col):
        """
        指定したセルの値を得る
        :param row: 行番号
        :param col: 列番号
        :return: セル値
        """
        return  self.__sheet.cell( row=row, column=col ).value

    def newSheet(self, sheetName=None ):
        """
        新しいシートを作成する
        :param sheetName: 作成するシート名（省略可:省略時sheet)
        重複している場合は、末尾に数字がついたシート名になる
        """
        self.__sheet = self.__book.create_sheet(sheetName)

    def numOfRow(self):
        """
        フォーカスしているシートの行数を得る
        :return: 行数
        """
        return self.__sheet.max_row

    def numOfCol(self):
        """
        フォーカスしているシートの列数を得る
        :return: 列数
        """
        return self.__sheet.max_column

    def setCell(self, row, col,  value):
        """
        指定したセルに値をセットする
        :param row: 行番号
        :param col: 列番号
        :param value: セットする値
        """
        self.__sheet.cell( row=row, column=col, value=value )

    def saveBook(self, bookName=None ):
        """
        ブックを保存する
        :param bookName: 保存するブック名（省略時上書き)
        :return:
        """
        if ( bookName == None ):
            bookName = self.__fileName

        self.__fileName = bookName
        self.__book.save(bookName)

    @classmethod
    def A1toR1C1( cls, a1Str ):
        """
        A1形式のセル指定をR1C1形式に返還する

        :param a1Str:A1形式のセル指定
        :return:[r, c] r=行番号,c列番号
        """
        col = re.search(r'^[A-Z,a-z]+', a1Str )             # 英大文字小文字を1回以上繰り返したものを抽出
        row = re.search(r'[0-9]+', a1Str )                   # 数字を1回以上繰り返したものを抽出

        if ( col == None ):                                 # 列番号指定が無ければ終わり
            return None

        colNum = cls.ColumnStrToNum(col.group(0))            # 列番号文字列を列番号に変換
        if (colNum==None):
            return None

        if (row == None):
            return None
        else:
            rowNum = int(row.group(0))

        return [rowNum, colNum]                           # 列番号は1から始めるので1オリジンに変換

    @classmethod
    def ColumnStrToNum( cls, colStr ):
        """
        A1形式の列指定をR1C1形式の列番号に

        :param colStr　列名文字列
        :return:列番号
        """
        col = re.search(r'^[A-Z,a-z]+', colStr )             # 英大文字小文字を1回以上繰り返したものを抽出

        if ( col == None ):                                 # 列番号指定が無ければ終わり
            return None
        else:
            column = [ ord(c) for c in col.group(0).upper()]    # カラム名を大文字にしてから文字コード列に変換

            # 文字コードを26進数文字として数値に変換
            colNum = 0
            for n in column:
                colNum = colNum*26 + (n-0x41+1)                 # 列番号は1から始めるので1オリジンに変換

        return colNum


###########################
#
#   テスト用
#
#
#r  >python ExcelMan.py
#       sheetList=
#       ['SheetNo1', 'SheetNo2', 'SheetNo3']
#       row= 28 colum= 11
#       <Cell 'SheetNo1'.B3>
#       Cell0,2
#
###########################
if __name__ == '__main__':
    print("####テストスタート#####")
    tergetObj = ExcelMan("sample.xlsx")             # "sample.xlsxファイルの管理オブジェクトを作る
    tergetObj.openBook()                               # "sample.xlsxファイルを開く
    sheetList = tergetObj.getSheetList()               # シートの一覧を得る
    print("sheetList=")
    print(sheetList)                                    #シートの一覧を表示
    tergetObj.openSheet(0)                              #先頭のシートを開く
    rows = tergetObj.numOfRow()                         #行数を得る
    cols = tergetObj.numOfCol()                         #列数を得る
    print( "row=", rows, "colum=", cols )           #行数と列数を表示
    cell = tergetObj.getCell( 3, 2 )                     #B3(R3C2)のセルを得る
    print(cell)                                         #B3のセルを表示
    value = tergetObj.getCellValue( 16,3 )              #C16のセルを表示
    print(value)
    tergetObj.setCell(27, 4, "CHANGED")                 # D27のセルの内容を変更
    tergetObj.newSheet()                                    #Sheetという名前のシートを追加
    tergetObj.newSheet("NewSheet")                      # NewSheetという名前のシートを追加
    tergetObj.saveBook("TestOut.xlsx")                  # "TestOut.xlsx"ファイルに保存

    out = ExcelMan.A1toR1C1("A1")
    print("A1=")
    print(out)
    out = ExcelMan.A1toR1C1("ZzZ65536")
    print("ZzZ65536=")
    print(out)
    out = ExcelMan.A1toR1C1("aa1")
    print("aa1=")
    print(out)
    out = ExcelMan.A1toR1C1("zz1")
    print("zz1=")
    print(out)
    out = ExcelMan.A1toR1C1("aaa1")
    print("aaa1=")
    print(out)
    out = ExcelMan.A1toR1C1("65536ZZZ")
    print("65536ZZZ=")
    print(out)
    out = ExcelMan.A1toR1C1("aa1a1a1a65536")
    print("aa1a1a1a65536=")
    print(out)
    out = ExcelMan.A1toR1C1("65536")
    print("65536=")
    print(out)
    out = ExcelMan.A1toR1C1("AAA")
    print("AAA=")
    print(out)

    print("####テスト終了#####")
